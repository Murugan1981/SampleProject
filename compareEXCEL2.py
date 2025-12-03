import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# === STEP 1: Paths ===
base_path = Path(__file__).resolve().parent.parent
input_folder = base_path / "shared" / "input" / "ExcelCompare"
output_file = input_folder / "PRD_UAT_EmbeddedDiff.xlsx"

file_prd = input_folder / "PRD.xlsm"
file_uat = input_folder / "UAT.xlsm"

# === STEP 2: Load Data ===
df_prd = pd.read_excel(file_prd, engine="openpyxl")
df_uat = pd.read_excel(file_uat, engine="openpyxl")

# Strip whitespace from headers
df_prd.columns = df_prd.columns.str.strip()
df_uat.columns = df_uat.columns.str.strip()

# === STEP 3: Validate PositionID ===
key_col = "PositionID"
if key_col not in df_prd.columns or key_col not in df_uat.columns:
    raise KeyError(f"❌ Missing '{key_col}' in one of the files.")

# === STEP 4: Set index and align ===
df_prd.set_index(key_col, inplace=True)
df_uat.set_index(key_col, inplace=True)

common_ids = df_prd.index.intersection(df_uat.index)

# Subset to common rows only
df_prd_common = df_prd.loc[common_ids]
df_uat_common = df_uat.loc[common_ids]

# === STEP 5: Build merged DataFrame ===
df_diff = pd.DataFrame(index=common_ids)

for col in df_prd.columns:
    prd_col = df_prd_common[col]
    uat_col = df_uat_common[col]

    def embed_diff(prd_val, uat_val):
        if pd.isna(prd_val) and pd.isna(uat_val):
            return ""
        elif prd_val != uat_val:
            return f"PRD: {prd_val} | UAT: {uat_val}"
        else:
            return prd_val

    df_diff[col] = [embed_diff(prd_col[i], uat_col[i]) for i in prd_col.index]

# Reset index to make PositionID a column again
df_diff.reset_index(inplace=True)

# === STEP 6: Write to Excel and highlight ===
wb = Workbook()
ws = wb.active
ws.title = "Comparison"

# Write rows
for row in dataframe_to_rows(df_diff, index=False, header=True):
    ws.append(row)

# Highlight yellow where embedded diff string exists
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

num_rows = ws.max_row
num_cols = ws.max_column

# Start from row 2 (skip header)
for r in range(2, num_rows + 1):
    for c in range(1, num_cols + 1):
        val = ws.cell(row=r, column=c).value
        if isinstance(val, str) and "PRD:" in val and "UAT:" in val and val != "":
            ws.cell(row=r, column=c).fill = highlight_fill

# === STEP 7: Save ===
wb.save(output_file)
print(f"✅ Final comparison saved to: {output_file}")
