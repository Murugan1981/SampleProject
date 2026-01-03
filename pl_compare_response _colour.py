import os
import json
import pandas as pd
from deepdiff import DeepDiff
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------- Load env vars --------
load_dotenv()
INPUT_XLSX = "shared/input/pl_responseFiles.xlsx"
APITESTDATA_FILE = "shared/input/ApiTestData.json"

# -------- Output File --------
OUTPUT_XLSX = "shared/reports/pl_comparison_result.xlsx"

# -------- Load System Name --------
with open(APITESTDATA_FILE, "r") as f:
    SYSTEM = json.load(f).get("System", "UNKNOWN")

RESPONSE_FOLDER = os.path.join("shared", "reports", SYSTEM)

# -------- Helper to Load JSON File --------
def load_json(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

# -------- Main Comparison --------
def compare_jsons(source_data, target_data):
    if source_data is None or target_data is None:
        return "NotMatch"
    diff = DeepDiff(source_data, target_data, ignore_order=True)
    return "Match" if not diff else "NotMatch"

def extract_snapshot(data):
    if data is None:
        return "EMPTY"
    if isinstance(data, list) and len(data) == 0:
        return "[]"
    if isinstance(data, dict) and len(data) == 0:
        return "{}"
    try:
        snippet = json.dumps(data, ensure_ascii=False)
        return snippet[:100] + ("..." if len(snippet) > 100 else "")
    except Exception:
        return "Invalid JSON"

def main():
    df = pd.read_excel(INPUT_XLSX)
    output_rows = []

    for _, row in df.iterrows():
        testcase_id = str(row.get("TestCaseID", "")).strip()
        tag = str(row.get("TagName", "")).strip()
        src_base = str(row.get("SourceBaseURL", "")).strip()
        tgt_base = str(row.get("TargetBaseURL", "")).strip()
        src_url = str(row.get("SourceRequestURL", "")).strip()
        tgt_url = str(row.get("TargetRequestURL", "")).strip()
        src_file = str(row.get("SourceResponse", "")).strip()
        tgt_file = str(row.get("TargetResponse", "")).strip()

        src_path = os.path.join(RESPONSE_FOLDER, src_file)
        tgt_path = os.path.join(RESPONSE_FOLDER, tgt_file)

        src_json = load_json(src_path)
        tgt_json = load_json(tgt_path)

        result = compare_jsons(src_json, tgt_json)

        src_snapshot = extract_snapshot(src_json) if result == "NotMatch" else ""
        tgt_snapshot = extract_snapshot(tgt_json) if result == "NotMatch" else ""

        output_rows.append({
            "TestCaseID": testcase_id,
            "TagName": tag,
            "SourceBaseURL": src_base,
            "TargetBaseURL": tgt_base,
            "SourceRequestURL": src_url,
            "TargetRequestURL": tgt_url,
            "SourceResponse": src_file,
            "TargetResponse": tgt_file,
            "ComparisonResult": result,
            "SourceSnapshot": src_snapshot,
            "TargetSnapshot": tgt_snapshot
        })

    df_out = pd.DataFrame(output_rows)
    df_out.to_excel(OUTPUT_XLSX, index=False)

    # -------- Highlight cells --------
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    blue_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        result_cell = row[8]  # ComparisonResult
        src_snap_cell = row[9]  # SourceSnapshot
        tgt_snap_cell = row[10] # TargetSnapshot

        if result_cell.value == "NotMatch":
            if src_snap_cell.value:
                src_snap_cell.fill = yellow_fill
            if tgt_snap_cell.value:
                tgt_snap_cell.fill = blue_fill

    wb.save(OUTPUT_XLSX)
    print(f"✅ Comparison complete. Output written to → {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
